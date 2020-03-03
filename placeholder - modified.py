# -*- coding: utf-8 -*-
"""
Created on Mon Oct  7 23:18:25 2019

@author: Abirmoy
"""

import xlrd
from openpyxl import load_workbook



# ENTER THE LOCATION OF THE FILE MODIFICATION WILL BE PERFORMED
workbook = xlrd.open_workbook(r"C:\Users\Python\source_m.xlsx")
book = load_workbook(r"C:\Users\\Python\source_m.xlsx")

worksheet = workbook.sheet_by_name('Sheet1')
sheet = book.active
num_rows = worksheet.nrows


def rows(y):
    '''
    Argument: This function takes integer as Row of an Exel file
    Returns: String
    Works for 5 column
    '''
    value1 = sheet.cell(y, 1).value # (row, column)
    value2 = sheet.cell(y, 2).value
    value3 = sheet.cell(y, 3).value
    try:
        if len(value3)==0:
            value3 = None
    except TypeError:
        value3 = None
        
    value4 = sheet.cell(y, 4).value
    value5 = sheet.cell(y, 5).value
    
    string = f"[!ab_{value1}_!H_{value2}_!cd_{value3}_!ef_{value4}_!g_{value5}]"
    string1 = string.replace('None', "''")
    return string1


for x in range(2, num_rows+1):
    
    sheet.cell(row = x, column = 7).value = rows(x) # MODIFIED INFORMATION WILL STORED AT 7TH COLUMN

    # PLEASE MODIFY THE LOCATION BELOW TO WITH DIRECTORY AND NAME FOR GENERATE NEW FILE    
    book.save(r"C:\\Abir\NewFile.xlsx")

print("Completed")
